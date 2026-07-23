# leads/views.py
from django.views.generic import TemplateView
from django.views.generic import DetailView
from django.utils.decorators import method_decorator
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import get_object_or_404, redirect
from django.shortcuts import render
from django.core.paginator import Paginator
from users.permissions import role_required
from django.db.models import Q
from django.utils import timezone
from .mixins import LeadOwnershipMixin
from .models import CoreLead, Notificacion
from users.models import CatUbicacion, CatEspecialidad, CatProducto, CatTitulo
from django.utils.timezone import localtime, now
from users.permissions import role_required


@method_decorator(role_required(['VENDEDOR']), name='dispatch')
class DashboardAgenteView(LoginRequiredMixin, LeadOwnershipMixin, TemplateView):
    template_name = 'dashboard_agente.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # 1. CAPTURAR LO QUE EL VENDEDOR QUIERE VER
        busqueda = self.request.GET.get('q', '').strip()
        filtro_rapido = self.request.GET.get('filtro', 'activos') # 'activos' por defecto
        producto_filtro = self.request.GET.get('producto', '').strip()
        
        #hoy = timezone.now().date()
        hoy = localtime(now()).date()   

        # 2. BASE DE SEGURIDAD: Solo los leads de ESTE vendedor
        qs = CoreLead.objects.filter(owner=self.request.user)

        # Filtrar por producto (familia EQUIPO) si se especifica
        if producto_filtro:
            qs = qs.filter(producto_cat__nombre=producto_filtro)

        # ---------------------------------------------------------
        # ESCENARIO A: EL FRANCOTIRADOR (Buscando un registro específico)
        # ---------------------------------------------------------
        if busqueda:
            # Si hay búsqueda, rompemos las reglas y buscamos en TODA su cartera histórica
            qs = qs.filter(
                Q(nombre_pila__icontains=busqueda) |
                Q(apellido_paterno__icontains=busqueda) |
                Q(apellido_materno__icontains=busqueda) |
                Q(phone_primary__icontains=busqueda) |
                Q(celular__icontains=busqueda) |
                Q(email__icontains=busqueda) |
                Q(clinica__nombre__icontains=busqueda)
            )
            
        # ---------------------------------------------------------
        # ESCENARIO B: LA RED DE ARRASTRE (Filtrando grupos diarios)
        # ---------------------------------------------------------
        else:
            if filtro_rapido == 'cierres':
                mes_actual = localtime(now())
                qs = qs.filter(
                    estatus='CLIENTE',
                    updated_at__year=mes_actual.year,
                    updated_at__month=mes_actual.month
                )
            else:
                # Regla INBOX ZERO: (Se mantiene para todo lo demás)
                qs = qs.exclude(estatus__in=['CLIENTE', 'NO_CIERRE']).exclude(plan='DESCARTADO')
                
                # Regla HIBERNACIÓN: (Se mantiene para todo lo demás)
                qs = qs.exclude(Q(plan='EN_ESPERA') & Q(next_action_date__gt=hoy))
    
                # Aplicar el botón que el vendedor haya presionado
                if filtro_rapido == 'hoy':
                    qs = qs.filter(next_action_date=hoy)
                elif filtro_rapido == 'frescos':
                    qs = qs.filter(estatus='PROSPECTO')
                elif filtro_rapido == 'leads':
                    qs = qs.filter(estatus='LEAD')
                elif filtro_rapido == 'calificados':
                    qs = qs.filter(estatus='LEAD_CALIFICADO')
                elif filtro_rapido == 'urgentes':
                    # Ejemplo: leads en SEGUIMIENTO que su fecha de acción ya se pasó
                    qs = qs.filter(plan='SEGUIMIENTO', next_action_date__lt=hoy)
                elif filtro_rapido == 'campanas':
                    evento_id = self.request.GET.get('evento_id')
                    if evento_id:
                        qs = qs.filter(eventos_asociados__evento_id=evento_id)
                    else:
                        qs = qs.filter(eventos_asociados__isnull=False).distinct()
                elif filtro_rapido == 'expos':
                    from .models import Evento
                    filtro_expos = self.request.GET.get('filtro_expos', 'activas')
                    qs_expos = Evento.objects.filter(vendedores_asignados=self.request.user, tipo='EXPO')
                    if filtro_expos == 'finalizadas':
                        qs_expos = qs_expos.filter(estatus='FINALIZADO')
                    else:
                        qs_expos = qs_expos.filter(estatus='ACTIVO')
                    context['expos_list'] = qs_expos.order_by('fecha_inicio')
                    context['filtro_expos'] = filtro_expos
                    qs = CoreLead.objects.none()
        
        # Ordenamos la consulta final con el criterio de prioritización (revisados hoy al final)
        from django.db.models import Case, When, Value, BooleanField
        qs = qs.annotate(
            revisado_hoy=Case(
                When(updated_at__date=hoy, then=Value(True)),
                default=Value(False),
                output_field=BooleanField()
            )
        ).order_by('revisado_hoy', '-updated_at')

        # Dividimos en bloques de 10 registros por página
        paginador = Paginator(qs, 7) 
        numero_pagina = self.request.GET.get('page')
        pagina_obj = paginador.get_page(numero_pagina)

        # 3. ENVIAR RESULTADOS AL HTML
        # OJO: Ahora mandamos la página actual, no toda la base de datos
        context['leads'] = pagina_obj 
        context['page_obj'] = pagina_obj # Lo mandamos también con este nombre por convención de Django
        
        context['busqueda_actual'] = busqueda
        context['filtro_actual'] = filtro_rapido
        context['producto_actual'] = producto_filtro
        
        context['total_activos'] = CoreLead.objects.filter(owner=self.request.user).exclude(estatus__in=['CLIENTE', 'NO_CIERRE']).exclude(plan='DESCARTADO').count()
        # --- LÍNEAS NUEVAS PARA EL MODAL DE ALTA RÁPIDA ---
        context['especialidades_list'] = CatEspecialidad.objects.filter(is_active=True).values_list('nombre', flat=True).order_by('nombre')
        context['productos_list'] = CatProducto.objects.filter(is_active=True, familia='EQUIPO').values_list('nombre', flat=True).order_by('nombre')
        context['ubicaciones_list'] = CatUbicacion.objects.filter(is_active=True).values_list('ciudad', flat=True).order_by('ciudad')
        context['titulos_list'] = CatTitulo.objects.filter(is_active=True).order_by('nombre')
        
        from .models import Evento
        context['active_expos'] = Evento.objects.filter(vendedores_asignados=self.request.user, tipo='EXPO', estatus='ACTIVO').order_by('nombre')

        
        # KPIs del Embudo de Ventas — Fila superior del Dashboard
        qs_owner = CoreLead.objects.filter(owner=self.request.user)
        if producto_filtro:
            qs_owner = qs_owner.filter(producto_cat__nombre=producto_filtro)
        mes_actual = localtime(now())

        context['total_activos']     = qs_owner.exclude(estatus__in=['CLIENTE', 'NO_CIERRE']).exclude(plan='DESCARTADO').count()
        context['total_prospectos']  = qs_owner.filter(estatus='PROSPECTO').count()
        context['total_leads']       = qs_owner.filter(estatus='LEAD').count()
        context['total_calificados'] = qs_owner.filter(estatus='LEAD_CALIFICADO').count()
        context['total_cierres_mes'] = qs_owner.filter(
            estatus='CLIENTE',
            es_historico=False,
            updated_at__year=mes_actual.year,
            updated_at__month=mes_actual.month
        ).count()

        return context

@method_decorator(role_required(['VENDEDOR']), name='dispatch')
class AgenteExpoCapturaView(LoginRequiredMixin, LeadOwnershipMixin, TemplateView):
    template_name = 'agente_expo_captura.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        evento_id = self.kwargs.get('evento_id')
        
        from .models import Evento
        evento_seleccionado = get_object_or_404(
            Evento, 
            id=evento_id, 
            vendedores_asignados=self.request.user, 
            tipo='EXPO'
        )
        
        # Filtramos los leads de este vendedor vinculados a este evento, ordenando prioritariamente
        from django.db.models import Case, When, Value, BooleanField
        from django.utils import timezone
        hoy_fecha = timezone.localdate()
        
        qs = CoreLead.objects.filter(
            owner=self.request.user,
            eventos_asociados__evento=evento_seleccionado
        ).annotate(
            revisado_hoy=Case(
                When(updated_at__date=hoy_fecha, then=Value(True)),
                default=Value(False),
                output_field=BooleanField()
            )
        ).order_by('revisado_hoy', '-updated_at')
        
        # Dividimos en bloques de 7 registros por página
        paginador = Paginator(qs, 7)
        numero_pagina = self.request.GET.get('page')
        pagina_obj = paginador.get_page(numero_pagina)
        
        context['evento_seleccionado'] = evento_seleccionado
        context['leads'] = pagina_obj
        context['page_obj'] = pagina_obj
        
        # Catálogos para el modal de alta rápida
        context['especialidades_list'] = CatEspecialidad.objects.filter(is_active=True).values_list('nombre', flat=True).order_by('nombre')
        context['productos_list'] = CatProducto.objects.filter(is_active=True, familia='EQUIPO').values_list('nombre', flat=True).order_by('nombre')
        context['ubicaciones_list'] = CatUbicacion.objects.filter(is_active=True).values_list('ciudad', flat=True).order_by('ciudad')
        context['titulos_list'] = CatTitulo.objects.filter(is_active=True).order_by('nombre')
        
        context['filtro_actual'] = 'expos'
        
        return context

@login_required
def agente_exportar_leads_view(request):
    """Exporta los leads del agente logueado a XLSX, respetando los filtros del dashboard."""
    busqueda      = request.GET.get('q', '').strip()
    filtro_rapido = request.GET.get('filtro', 'activos')
    producto_filtro = request.GET.get('producto', '').strip()
    hoy           = localtime(now()).date()

    qs = CoreLead.objects.filter(owner=request.user)

    if producto_filtro:
        qs = qs.filter(producto_cat__nombre=producto_filtro)

    if busqueda:
        qs = qs.filter(
            Q(nombre_pila__icontains=busqueda) |
            Q(apellido_paterno__icontains=busqueda) |
            Q(apellido_materno__icontains=busqueda) |
            Q(phone_primary__icontains=busqueda) |
            Q(celular__icontains=busqueda) |
            Q(email__icontains=busqueda) |
            Q(clinica__nombre__icontains=busqueda)
        )
    else:
        if filtro_rapido == 'cierres':
            mes_actual = localtime(now())
            qs = qs.filter(
                estatus='CLIENTE',
                updated_at__year=mes_actual.year,
                updated_at__month=mes_actual.month
            )
        else:
            qs = qs.exclude(estatus__in=['CLIENTE', 'NO_CIERRE']).exclude(plan='DESCARTADO')
            qs = qs.exclude(Q(plan='EN_ESPERA') & Q(next_action_date__gt=hoy))
            if filtro_rapido == 'hoy':
                qs = qs.filter(next_action_date=hoy)
            elif filtro_rapido == 'frescos':
                qs = qs.filter(estatus='PROSPECTO')
            elif filtro_rapido == 'leads':
                qs = qs.filter(estatus='LEAD')
            elif filtro_rapido == 'calificados':
                qs = qs.filter(estatus='LEAD_CALIFICADO')
            elif filtro_rapido == 'urgentes':
                qs = qs.filter(plan='SEGUIMIENTO', next_action_date__lt=hoy)

    qs = qs.order_by('-updated_at')
    return generar_respuesta_xlsx(qs, f"mis_leads_{request.user.username}.xlsx")


@login_required
def agente_exportar_talleres_view(request):
    """
    Exporta los prospectos/clientes de talleres o campañas del vendedor a XLSX.
    Soporta:
    - Reporte Parcial: enviando ?evento_id=<id> (exporta solo ese evento)
    - Reporte Total: sin evento_id, exporta todos los eventos del tipo según ?tab=campanas|talleres
      respetando el filtro de tiempo ?filtro_evento=proximos_7|este_mes|finalizados|todos
    """
    from django.db.models import Q
    from django.utils import timezone
    from django.http import HttpResponse
    import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from .models import Evento, CoreLead, LeadEvento
    
    user = request.user
    evento_id = request.GET.get('evento_id')
    tab = request.GET.get('tab', 'talleres')
    filtro_evento = request.GET.get('filtro_evento', 'todos')
    
    # Determinar tipo de evento
    tipo_evento = 'TALLER'
    if tab == 'campanas':
        tipo_evento = 'CAMPAÑA'
        
    qs_eventos = Evento.objects.filter(vendedores_asignados=user)
    
    # Si viene evento_id, es reporte parcial
    if evento_id:
        evento_seleccionado = get_object_or_404(Evento, id=evento_id, vendedores_asignados=user)
        eventos = [evento_seleccionado]
        nombre_archivo = f"reporte_parcial_{evento_seleccionado.nombre}.xlsx"
    else:
        # Reporte total
        qs_eventos = qs_eventos.filter(tipo=tipo_evento)
        hoy = timezone.now().date()
        
        if filtro_evento == 'proximos_7':
            limite = hoy + datetime.timedelta(days=7)
            qs_eventos = qs_eventos.filter(estatus='ACTIVO', fecha_inicio__gte=hoy, fecha_inicio__lte=limite).order_by('fecha_inicio')
        elif filtro_evento == 'este_mes':
            qs_eventos = qs_eventos.filter(estatus='ACTIVO', fecha_inicio__year=hoy.year, fecha_inicio__month=hoy.month).order_by('fecha_inicio')
        elif filtro_evento == 'finalizados':
            qs_eventos = qs_eventos.filter(estatus='FINALIZADO').order_by('-fecha_inicio')
        else:
            qs_eventos = qs_eventos.filter(estatus='ACTIVO').order_by('fecha_inicio')
            
        eventos = list(qs_eventos)
        tipo_label = "talleres" if tipo_evento == 'TALLER' else "campanas"
        nombre_archivo = f"reporte_total_{tipo_label}_{filtro_evento}.xlsx"

    # Preparar el Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Talleres" if tipo_evento == 'TALLER' else "Reporte Campañas"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1E293B") # Navy
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    bold_font = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )
    
    fill_vinculado = PatternFill("solid", fgColor="E2F0D9") # Light Green
    fill_candidato = PatternFill("solid", fgColor="F2F2F2") # Light Gray
    
    HEADERS = [
        "Tipo de Evento", "Nombre de Evento/Taller", "Línea de Producto", 
        "Fecha Inicio", "Fecha Fin", "Lugar/Sede", "Relación con Cliente",
        "Nombre Completo Cliente", "Teléfono", "Celular", "Email",
        "Estatus Lead", "Especialidad", "Ciudad", "Último Abordaje",
        "Fecha Vinculación", "Comentarios de Vinculación"
    ]
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    ws.row_dimensions[1].height = 28

    # Recopilar todos los leads para los eventos
    row_idx = 2
    for ev in eventos:
        qs_base = CoreLead.objects.filter(owner=user)
        linea = ev.linea_producto
        ids_permitidos = obtener_especialidades_permitidas(linea)
        
        MAP_LINEA_PRODUCTO = {
            'SPORT': 'Sport',
            'PET': 'Pet',
            'DENTAL': 'Dental',
            'PODOLOGICO': 'Podológico',
            'BEAUTY': 'Beauty'
        }
        keyword_producto = MAP_LINEA_PRODUCTO.get(linea)
        
        q_linea = Q(especialidad_cat_id__in=ids_permitidos)
        if keyword_producto:
            q_linea &= Q(producto_cat__nombre__icontains=keyword_producto)

        ciudades_objetivo = ev.ciudades_objetivo.all()
        if ciudades_objetivo.exists():
            q_filter = Q(ubicacion__in=ciudades_objetivo)
            q_filter &= q_linea
            qs_prospectos = qs_base.filter(
                Q(eventos_asociados__evento=ev) |
                q_filter
            ).filter(estatus='CLIENTE').distinct()
        else:
            q_filter = q_linea
            qs_prospectos = qs_base.filter(
                Q(eventos_asociados__evento=ev) |
                q_filter
            ).filter(estatus='CLIENTE').distinct()

        # Excluir los leads que ya fueron abordados (tienen oportunidad 360 reciente)
        fecha_margen = ev.fecha_inicio - datetime.timedelta(days=15)
        qs_prospectos = qs_prospectos.exclude(
            compras_extra__vendedor=user,
            compras_extra__fecha_venta__gte=fecha_margen
        ).order_by('-updated_at')
        
        # Obtener vinculaciones manuales
        vinculos_qs = LeadEvento.objects.filter(evento=ev, lead__in=qs_prospectos).select_related('lead')
        vinculos_map = {v.lead_id: v for v in vinculos_qs}
        
        for lead in qs_prospectos:
            vinculo = vinculos_map.get(lead.id)
            relacion = "VINCULADO" if vinculo else "CANDIDATO POR PERFIL"
            fecha_vinc_str = timezone.localtime(vinculo.fecha_vinculacion).strftime("%d/%m/%Y %H:%M") if vinculo else ""
            comentarios_vinc = vinculo.comentarios or "" if vinculo else ""
            
            ws.append([
                ev.get_tipo_display(),
                ev.nombre,
                ev.get_linea_producto_display(),
                ev.fecha_inicio.strftime("%d/%m/%Y") if ev.fecha_inicio else "",
                ev.fecha_fin.strftime("%d/%m/%Y") if ev.fecha_fin else "",
                ev.lugar or "",
                relacion,
                lead.nombre_completo_mdm,
                lead.phone_primary or "",
                lead.celular or "",
                lead.email or "",
                lead.estatus,
                lead.especialidad_cat.nombre if lead.especialidad_cat else "",
                lead.ubicacion.ciudad if lead.ubicacion else "",
                lead.updated_at.strftime("%d/%m/%Y") if lead.updated_at else "",
                fecha_vinc_str,
                comentarios_vinc
            ])
            
            # Estilos de fila
            fill_to_use = fill_vinculado if vinculo else fill_candidato
            ws.cell(row=row_idx, column=7).font = bold_font
            ws.cell(row=row_idx, column=7).fill = fill_to_use
            
            for col_idx in range(1, len(HEADERS) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                
            row_idx += 1

    # Ajustar anchos
    ANCHOS = [18, 30, 18, 14, 14, 25, 22, 32, 14, 14, 28, 14, 24, 20, 16, 20, 35]
    for i, ancho in enumerate(ANCHOS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = ancho

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    # Reemplazar espacios y caracteres raros en el nombre de archivo
    import re
    nombre_archivo_clean = re.sub(r'[^\w\.-]', '_', nombre_archivo)
    response["Content-Disposition"] = f'attachment; filename="{nombre_archivo_clean}"'
    wb.save(response)
    return response


@method_decorator(role_required(['VENDEDOR']), name='dispatch')
class Ventas360View(LoginRequiredMixin, TemplateView):
    template_name = 'ventas_360.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        qs_base = CoreLead.objects.filter(owner=self.request.user)
        
        from .models import Evento, VentaTransaccional
        
        # 1. Oportunidades 360 (Ventas Transaccionales)
        qs_oportunidades = VentaTransaccional.objects.filter(vendedor=self.request.user).select_related('lead', 'producto').prefetch_related('lead__eventos_asociados').order_by('-fecha_venta')
        estatus_oportunidad = self.request.GET.get('estatus')
        if estatus_oportunidad == 'gestion':
            qs_oportunidades = qs_oportunidades.filter(estatus__in=['PENDIENTE', 'EN_GESTION'])
        elif estatus_oportunidad == 'concretadas':
            qs_oportunidades = qs_oportunidades.filter(estatus='CONCRETADO')
            
        paginator_oportunidades = Paginator(qs_oportunidades, 10)
        page_oportunidades = self.request.GET.get('page_oportunidades')
        context['oportunidades_360'] = paginator_oportunidades.get_page(page_oportunidades)
        
        # 2. Históricos
        qs_historicos = qs_base.filter(es_historico=True).order_by('-updated_at')
        paginator_historicos = Paginator(qs_historicos, 10)
        page_historicos = self.request.GET.get('page_historicos')
        context['historicos'] = paginator_historicos.get_page(page_historicos)
        
        # Pestaña activa por defecto
        context['active_tab'] = self.request.GET.get('tab', 'mis_clientes')
        
        # 3. Mis Clientes (Leads no históricos con estatus CLIENTE)
        qs_mis_clientes = qs_base.filter(estatus='CLIENTE', es_historico=False).order_by('-updated_at')
        paginator_mis_clientes = Paginator(qs_mis_clientes, 10)
        page_mis_clientes = self.request.GET.get('page_mis_clientes')
        context['mis_clientes_list'] = paginator_mis_clientes.get_page(page_mis_clientes)
        
        # 3. Campañas y Eventos (Separados)
        filtro_evento = self.request.GET.get('filtro_evento', 'todos')
        context['filtro_evento'] = filtro_evento
        
        from django.utils import timezone
        import datetime
        hoy = timezone.now().date()
        
        qs_eventos = Evento.objects.filter(vendedores_asignados=self.request.user)
        if context['active_tab'] == 'campanas':
            qs_eventos = qs_eventos.filter(tipo='CAMPAÑA')
        elif context['active_tab'] == 'talleres':
            qs_eventos = qs_eventos.filter(tipo='TALLER')
        else:
            qs_eventos = qs_eventos.exclude(tipo='EXPO')
        
        if filtro_evento == 'proximos_7':
            limite = hoy + datetime.timedelta(days=7)
            qs_eventos = qs_eventos.filter(estatus='ACTIVO', fecha_inicio__gte=hoy, fecha_inicio__lte=limite).order_by('fecha_inicio')
        elif filtro_evento == 'este_mes':
            qs_eventos = qs_eventos.filter(estatus='ACTIVO', fecha_inicio__year=hoy.year, fecha_inicio__month=hoy.month).order_by('fecha_inicio')
        elif filtro_evento == 'finalizados':
            qs_eventos = qs_eventos.filter(estatus='FINALIZADO').order_by('-fecha_inicio')
        else:
            qs_eventos = qs_eventos.filter(estatus='ACTIVO').order_by('fecha_inicio')
            
        context['eventos_activos'] = qs_eventos

        # Manejo de Opción A: Prospectos in-place
        evento_id = self.request.GET.get('evento_id')
        if evento_id:
            evento_seleccionado = Evento.objects.filter(id=evento_id).exclude(tipo='EXPO').first()
            if evento_seleccionado:
                context['evento_seleccionado'] = evento_seleccionado
                from django.db.models import Q
                
                linea = evento_seleccionado.linea_producto
                ids_permitidos = obtener_especialidades_permitidas(linea)
                
                # Mapeo de Línea de Producto a palabras clave de Producto
                MAP_LINEA_PRODUCTO = {
                    'SPORT': 'Sport',
                    'PET': 'Pet',
                    'DENTAL': 'Dental',
                    'PODOLOGICO': 'Podológico',
                    'BEAUTY': 'Beauty'
                }
                keyword_producto = MAP_LINEA_PRODUCTO.get(linea)
                
                q_linea = Q(especialidad_cat_id__in=ids_permitidos)
                if keyword_producto:
                    q_linea &= Q(producto_cat__nombre__icontains=keyword_producto)

                # Buscar leads asignados manualmente O clientes en las ciudades objetivo que cumplan con la línea de producto/especialidad, con estatus CLIENTE únicamente
                ciudades_objetivo = evento_seleccionado.ciudades_objetivo.all()
                if ciudades_objetivo.exists():
                    q_filter = Q(ubicacion__in=ciudades_objetivo)
                    q_filter &= q_linea
                    qs_prospectos = qs_base.filter(
                        Q(eventos_asociados__evento=evento_seleccionado) |
                        q_filter
                    ).filter(estatus='CLIENTE').distinct()
                else:
                    q_filter = q_linea
                    qs_prospectos = qs_base.filter(
                        Q(eventos_asociados__evento=evento_seleccionado) |
                        q_filter
                    ).filter(estatus='CLIENTE').distinct()

                # Excluir los leads que ya fueron abordados (tienen oportunidad 360 reciente)
                import datetime
                fecha_margen = evento_seleccionado.fecha_inicio - datetime.timedelta(days=15)
                qs_prospectos = qs_prospectos.exclude(
                    compras_extra__vendedor=self.request.user,
                    compras_extra__fecha_venta__gte=fecha_margen
                ).order_by('-updated_at')
                
                paginator_prospectos = Paginator(qs_prospectos, 10)
                page_prospectos = self.request.GET.get('page_prospectos')
                context['prospectos_campana'] = paginator_prospectos.get_page(page_prospectos)

        # KPIs del resumen superior — Ventas 360
        context['kpi_mis_clientes']        = qs_base.filter(estatus='CLIENTE', es_historico=False).count()
        
        qs_oport_all = VentaTransaccional.objects.filter(vendedor=self.request.user)
        context['kpi_en_gestion']          = qs_oport_all.filter(estatus__in=['PENDIENTE', 'EN_GESTION']).count()
        context['kpi_concretadas']         = qs_oport_all.filter(estatus='CONCRETADO').count()
        context['kpi_historicos']          = qs_base.filter(es_historico=True).count()
        context['kpi_campanas_activas']    = Evento.objects.filter(
            vendedores_asignados=self.request.user, estatus='ACTIVO', tipo='CAMPAÑA'
        ).count()
        context['kpi_talleres_activos']    = Evento.objects.filter(
            vendedores_asignados=self.request.user, estatus='ACTIVO', tipo='TALLER'
        ).count()

        return context

@method_decorator(role_required(['VENDEDOR', 'DIRECTOR', 'ADMIN']), name='dispatch')
class IngestaMasivaView(LoginRequiredMixin, LeadOwnershipMixin, TemplateView):
    template_name = 'ingesta_masiva.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from .models import CatUbicacion, CatEspecialidad, CatProducto
        context['ubicaciones_list'] = CatUbicacion.objects.all().order_by('ciudad')
        context['especialidades_list'] = CatEspecialidad.objects.all().order_by('nombre')
        context['productos_list'] = CatProducto.objects.all().order_by('nombre')
        return context

@method_decorator(role_required(['VENDEDOR', 'DIRECTOR', 'ADMIN']), name='dispatch')
class AltaIndividualView(LoginRequiredMixin, LeadOwnershipMixin, TemplateView):
    template_name = 'alta_individual.html'

import os
import uuid
import pandas as pd
from django.conf import settings
from django.core.files.storage import FileSystemStorage
from django.contrib import messages
from leads.parser_service import orquestar_ingesta_historica

@method_decorator(role_required(['DIRECTOR', 'ADMIN']), name='dispatch')
class IngestaHistoricaView(LoginRequiredMixin, TemplateView):
    template_name = 'leads/ingesta_historica.html'

    def get_context_data(self, **kwargs):
        return super().get_context_data(**kwargs)

    def _get_fs_storage(self):
        """Helper para configurar el storage temporal de archivos de ingesta"""
        temp_dir = os.path.join(settings.MEDIA_ROOT, 'temp_ingesta')
        os.makedirs(temp_dir, exist_ok=True)
        return FileSystemStorage(location=temp_dir)

    def _procesar_archivo_pandas(self, fs, file_uuid):
        """Helper para leer el archivo guardado y devolver filas_data limpias"""
        file_path = fs.path(file_uuid)
        ext = os.path.splitext(file_uuid)[1].lower()
        if ext in ['.xls', '.xlsx']:
            df = pd.read_excel(file_path)
        else:
            df = pd.read_csv(file_path)

        # Sanitización inicial requerida (Fase 2)
        df = df.fillna('')
        df.columns = [str(c).strip().lower().replace(' ', '_') for c in df.columns]
        return df.to_dict('records')

    def post(self, request, *args, **kwargs):
        action = request.POST.get('action')
        context = self.get_context_data()
        fs = self._get_fs_storage()

        # Si la acción está vacía (Paso 1) o es 'simulate', entra aquí:
        if not action or action == 'simulate':
            archivo = request.FILES.get('archivo_historico')
            if not archivo:
                messages.error(request, "Debes adjuntar un archivo (CSV o Excel).")
                return render(request, self.template_name, context)

            ext = os.path.splitext(archivo.name)[1].lower()
            if ext not in ['.csv', '.xls', '.xlsx']:
                messages.error(request, "Formato no soportado. Usa CSV o Excel.")
                return render(request, self.template_name, context)

            try:
                # 1. Guardar temporalmente para persistencia en Paso 2
                file_uuid = f"{uuid.uuid4().hex}{ext}"
                fs.save(file_uuid, archivo)
                
                # 2. Leer y serializar con Pandas
                filas_data = self._procesar_archivo_pandas(fs, file_uuid)

                if not filas_data:
                    fs.delete(file_uuid)
                    messages.warning(request, "El archivo está vacío.")
                    return render(request, self.template_name, context)

                # 3. Orquestar SIMULACIÓN (dry_run=True)
                reporte = orquestar_ingesta_historica(
                    filas_data=filas_data,
                    admin_user=request.user,
                    dry_run=True
                )

                # 4. Enviar reporte al frontend
                context['reporte'] = reporte
                context['file_uuid'] = file_uuid
                context['simulacion_activa'] = True
                
                if reporte.get("errores_criticos", 0) > 0:
                    messages.warning(request, f"Revisión: {reporte['errores_criticos']} fila(s) con errores críticos omitidas.")

            except Exception as e:
                # Limpiar si falla la lectura inicial
                if 'file_uuid' in locals() and fs.exists(file_uuid):
                    fs.delete(file_uuid)
                messages.error(request, f"Error al procesar la simulación: {str(e)}")
            
            return render(request, self.template_name, context)

        elif action == 'commit':
            file_uuid = request.POST.get('file_uuid')
            if not file_uuid or not fs.exists(file_uuid):
                messages.error(request, "Sesión de simulacro expirada o archivo no encontrado. Repite el proceso.")
                return redirect('director_ingesta')

            try:
                # 1. Re-leer el archivo limpiado
                filas_data = self._procesar_archivo_pandas(fs, file_uuid)

                # 2. Orquestar INYECCIÓN REAL (dry_run=False)
                reporte = orquestar_ingesta_historica(
                    filas_data=filas_data,
                    admin_user=request.user,
                    dry_run=False
                )

                messages.success(
                    request, 
                    f"¡Migración Histórica completada! Se crearon {reporte['clinicas_identificadas']} clínicas y {reporte['individuos_atomizados']} individuos."
                )
            except Exception as e:
                messages.error(request, f"Error durante la inyección: {str(e)}")
            finally:
                # 3. Siempre eliminar el archivo temporal
                if fs.exists(file_uuid):
                    fs.delete(file_uuid)
                
            return redirect('director_dashboard')

        elif action == 'cancel':
            file_uuid = request.POST.get('file_uuid')
            if file_uuid and fs.exists(file_uuid):
                fs.delete(file_uuid)
            messages.info(request, "Proceso de ingesta cancelado de forma segura.")
            return redirect('director_ingesta')

        return render(request, self.template_name, context)

from leads.models import LeadStaging
from django.views.generic import ListView

@method_decorator(role_required(['DIRECTOR', 'ADMIN']), name='dispatch')
class ListaStagingView(LoginRequiredMixin, ListView):
    model = LeadStaging
    template_name = 'leads/staging_list.html'
    context_object_name = 'leads_staging'
    paginate_by = 50

    def get_queryset(self):
        # Solo mostrar los pendientes históricos
        return LeadStaging.objects.filter(estatus='PENDIENTE', origen='HISTORICO').order_by('created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['total_pendientes'] = LeadStaging.objects.filter(estatus='PENDIENTE', origen='HISTORICO').count()
        return context

@method_decorator(role_required(['DIRECTOR', 'ADMIN']), name='dispatch')
class ProcesarStagingView(LoginRequiredMixin, DetailView):
    model = LeadStaging
    template_name = 'leads/staging_procesar.html'
    context_object_name = 'staging_lead'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from django.contrib.auth import get_user_model
        User = get_user_model()
        context['vendedores_list'] = User.objects.filter(is_active=True, is_superuser=False).order_by('username')
        context['especialidades_list'] = CatEspecialidad.objects.filter(is_active=True).order_by('nombre')
        context['ubicaciones_list'] = CatUbicacion.objects.filter(is_active=True).order_by('ciudad')
        context['titulos_list'] = CatTitulo.objects.filter(is_active=True).order_by('nombre')
        context['productos_list'] = CatProducto.objects.filter(is_active=True, familia='EQUIPO').order_by('nombre')
        # Buscamos cuántos quedan para el badge superior (Históricos)
        context['restantes'] = LeadStaging.objects.filter(estatus='PENDIENTE', origen='HISTORICO').count()
        return context

    def post(self, request, *args, **kwargs):
        staging_lead = self.get_object()
        action = request.POST.get('action')

        if action == 'descartar':
            staging_lead.estatus = 'DESCARTADO'
            staging_lead.save()
            messages.info(request, "Registro descartado exitosamente.")
            return self._redirect_to_next()

        elif action == 'guardar':
            # Extraer IDs y tipos
            vendedor_id = request.POST.get('vendedor_id')
            tipo_entidad = request.POST.get('tipo_entidad', 'INDIVIDUAL')
            especialidad_id = request.POST.get('especialidad_id')
            ubicacion_id = request.POST.get('ubicacion_id')
            producto_id = request.POST.get('producto_id')

            telefono = str(request.POST.get('telefono', '')).strip()
            celular = str(request.POST.get('celular', '')).strip()
            email = str(request.POST.get('email', '')).strip()
            direccion = str(request.POST.get('direccion_completa', '')).strip()

            titulo_id = request.POST.get('titulo_cortesia')
            nombre_pila = str(request.POST.get('nombre_pila', '')).strip()
            apellido_paterno = str(request.POST.get('apellido_paterno', '')).strip()
            apellido_materno = str(request.POST.get('apellido_materno', '')).strip()

            vendedor_historico = staging_lead.datos_crudos.get('vendedor_historico', 'Desconocido')
            notas_originales = str(request.POST.get('notas', staging_lead.datos_crudos.get('notas', ''))).strip()

            # --- OBTENER OBJETOS DE CATÁLOGOS ---
            titulo_obj = CatTitulo.objects.filter(id=titulo_id).first() if titulo_id else None
            especialidad_obj = CatEspecialidad.objects.filter(id=especialidad_id).first() if especialidad_id else None
            ubicacion_obj = CatUbicacion.objects.filter(id=ubicacion_id).first() if ubicacion_id else None
            producto_obj = CatProducto.objects.filter(id=producto_id).first() if producto_id else None

            if not all([especialidad_obj, ubicacion_obj, producto_obj]):
                messages.error(request, "Especialidad, Ubicación y Producto de Interés son campos obligatorios.")
                return redirect('staging_procesar', pk=staging_lead.pk)

            # --- 1. PREPARACIÓN MDM ---
            datos_dict = {
                'tipo_entidad': tipo_entidad,
                'nombre_pila': nombre_pila,
                'apellido_paterno': apellido_paterno,
                'apellido_materno': apellido_materno,
                'telefono': telefono,
                'especialidad_obj': especialidad_obj,
                'ubicacion_obj': ubicacion_obj
            }

            # --- 2. ADUANA CENTRAL (MDM) ---
            try:
                from leads.services.mdm_service import resolver_identidad
                instancia_mdm, telefono_alternativo = resolver_identidad(datos_dict)
            except ValueError as e:
                # BLOQUEO MDM: Colisión Detectada
                messages.error(request, str(e))
                return redirect('staging_procesar', pk=staging_lead.pk)

            # --- ASIGNACIÓN DE DUEÑO (VENDEDOR) ---
            from django.contrib.auth import get_user_model
            from django.db.models import Count
            User = get_user_model()

            if not vendedor_id:
                vendedor_asignado = User.objects.filter(
                    is_active=True, is_superuser=False
                ).annotate(Count('leads')).order_by('leads__count').first()
            else:
                vendedor_asignado = get_object_or_404(User, id=vendedor_id)

            if not vendedor_asignado:
                 messages.error(request, "No hay vendedores activos en el sistema para asignar el prospecto.")
                 return redirect('staging_procesar', pk=staging_lead.pk)

            # Preparar notas históricas
            texto_nota = f"[QUIRÓFANO] Inyectado. Vendedor Orig: {vendedor_historico} | Notas: {notas_originales}"
            if telefono_alternativo:
                texto_nota += f" | MDM: Intento de inyección con teléfono alternativo: {telefono_alternativo}"

            notas_historicas = {
                "notas": [{
                    "tipo": "sistema",
                    "contenido": texto_nota,
                    "fecha": localtime(now()).isoformat(),
                    "usuario": request.user.id
                }],
                "columnas_excel_historicas": staging_lead.datos_crudos
            }

            # --- 3. INYECCIÓN O FUSIÓN SILENCIOSA ---
            try:
                if tipo_entidad == 'CORPORATIVO':
                    CoreLead.objects.create(
                        owner=vendedor_asignado,
                        ubicacion=ubicacion_obj,
                        estatus='CLIENTE',  
                        phone_primary=telefono if not telefono_alternativo else instancia_mdm.telefono_master,
                        celular=celular[:15],
                        email=email,
                        direccion_completa=direccion[:255],
                        nombre_pila=nombre_pila[:100],
                        clinica=instancia_mdm,
                        especialidad_cat=especialidad_obj,
                        producto_cat=producto_obj,
                        es_historico=True,
                        notas_variadas=notas_historicas,
                    )
                    messages.success(request, f"¡Inyectado exitosamente! Vinculado a Clínica Corporativa '{instancia_mdm.nombre}'. Asignado a {vendedor_asignado.username}.")

                else: # INDIVIDUAL
                    if instancia_mdm is not None:
                        # FUSIÓN: El registro ya existía en BD.
                        if not isinstance(instancia_mdm.notas_variadas, dict):
                            instancia_mdm.notas_variadas = {"notas": [], "columnas_excel_historicas": {}}
                        if "notas" not in instancia_mdm.notas_variadas:
                            instancia_mdm.notas_variadas["notas"] = []
                            
                        nota_fusion = {
                            "tipo": "sistema",
                            "contenido": f"MDM Quirófano: Fusión confirmada. Notas originales: {notas_originales}. Teléfono extra detectado: {telefono_alternativo or telefono}",
                            "fecha": localtime(now()).isoformat(),
                            "usuario": request.user.id
                        }
                        instancia_mdm.notas_variadas["notas"].append(nota_fusion)
                        instancia_mdm.save()
                        messages.success(request, "Registro fusionado/actualizado exitosamente por MDM (Doctor ya existía).")
                    else:
                        # NUEVO: Creación desde Quirófano desde cero
                        CoreLead.objects.create(
                            owner=vendedor_asignado,
                            ubicacion=ubicacion_obj,
                            estatus='CLIENTE',  
                            phone_primary=telefono,
                            celular=celular[:15],
                            email=email,
                            direccion_completa=direccion[:255],
                            titulo_cortesia=titulo_obj,
                            nombre_pila=nombre_pila[:100],
                            apellido_paterno=apellido_paterno[:100],
                            apellido_materno=apellido_materno[:100],
                            especialidad_cat=especialidad_obj,
                            producto_cat=producto_obj,
                            es_historico=True,
                            notas_variadas=notas_historicas,
                        )
                        messages.success(request, f"¡Inyectado exitosamente! Asignado a {vendedor_asignado.username}.")
                
                # Independientemente de Fusión o Nueva Inyección, purgar de Quirófano
                staging_lead.estatus = 'RESUELTO'
                staging_lead.save()
                
                return self._redirect_to_next()
                
            except Exception as e:
                messages.error(request, f"Error interno al guardar en BD: {str(e)}")
                return redirect('staging_procesar', pk=staging_lead.pk)

        return redirect('staging_procesar', pk=staging_lead.pk)

    def _redirect_to_next(self):
        # Busca el siguiente pendiente cronológicamente
        siguiente = LeadStaging.objects.filter(estatus='PENDIENTE').order_by('created_at').first()
        if siguiente:
            return redirect('staging_procesar', pk=siguiente.pk)
        messages.success(self.request, "¡Felicidades! Se ha vaciado la cola del Quirófano.")
        return redirect('staging_list')


from django.views.generic import View

@method_decorator(role_required(['DIRECTOR', 'ADMIN']), name='dispatch')
class IngestaHistoricaExpressView(LoginRequiredMixin, View):
    template_name = 'leads/ingesta_express.html'

    def get(self, request, *args, **kwargs):
        from django.contrib.auth import get_user_model
        User = get_user_model()
        context = {
            'vendedores_list': User.objects.filter(is_active=True, is_superuser=False).order_by('username'),
            'especialidades_list': CatEspecialidad.objects.filter(is_active=True).order_by('nombre'),
            'ubicaciones_list': CatUbicacion.objects.filter(is_active=True).order_by('ciudad'),
            'titulos_list': CatTitulo.objects.filter(is_active=True).order_by('nombre'),
            'productos_list': CatProducto.objects.filter(is_active=True, familia='EQUIPO').order_by('nombre'),
        }
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        vendedor_id = request.POST.get('vendedor_id')
        tipo_entidad = request.POST.get('tipo_entidad', 'INDIVIDUAL')
        especialidad_id = request.POST.get('especialidad_id')
        ubicacion_id = request.POST.get('ubicacion_id')
        producto_id = request.POST.get('producto_id')

        telefono = str(request.POST.get('telefono', '')).strip()
        celular = str(request.POST.get('celular', '')).strip()
        email = str(request.POST.get('email', '')).strip()
        
        titulo_id = request.POST.get('titulo_cortesia')
        nombre_pila = str(request.POST.get('nombre_pila', '')).strip()
        apellido_paterno = str(request.POST.get('apellido_paterno', '')).strip()
        apellido_materno = str(request.POST.get('apellido_materno', '')).strip()

        # Validación de campos y catálogos obligatorios
        titulo_obj = CatTitulo.objects.filter(id=titulo_id).first() if titulo_id else None
        especialidad_obj = CatEspecialidad.objects.filter(id=especialidad_id).first() if especialidad_id else None
        ubicacion_obj = CatUbicacion.objects.filter(id=ubicacion_id).first() if ubicacion_id else None
        producto_obj = CatProducto.objects.filter(id=producto_id).first() if producto_id else None

        if not all([especialidad_obj, ubicacion_obj, producto_obj]):
            messages.error(request, "Especialidad, Ubicación y Producto de Interés son campos obligatorios.")
            return redirect('ingesta_express')

        # --- 1. PREPARACIÓN MDM ---
        datos_dict = {
            'tipo_entidad': tipo_entidad,
            'nombre_pila': nombre_pila,
            'apellido_paterno': apellido_paterno,
            'apellido_materno': apellido_materno,
            'telefono': telefono,
            'especialidad_obj': especialidad_obj,
            'ubicacion_obj': ubicacion_obj
        }

        # --- 2. ADUANA CENTRAL (MDM) ---
        try:
            from leads.services.mdm_service import resolver_identidad
            instancia_mdm, telefono_alternativo = resolver_identidad(datos_dict)
        except ValueError as e:
            # BLOQUEO MDM: Colisión de Identidad Detectada
            messages.error(request, str(e))
            return redirect('ingesta_express')

        # --- ASIGNACIÓN DE DUEÑO (VENDEDOR) ---
        from django.contrib.auth import get_user_model
        from django.db.models import Count
        User = get_user_model()

        if not vendedor_id:
            vendedor_asignado = User.objects.filter(
                is_active=True, is_superuser=False
            ).annotate(Count('leads')).order_by('leads__count').first()
        else:
            vendedor_asignado = get_object_or_404(User, id=vendedor_id)

        if not vendedor_asignado:
            messages.error(request, "No hay vendedores activos en el sistema para asignar.")
            return redirect('ingesta_express')

        # Preparar notas históricas compartidas
        texto_nota = "[INGRESO EXPRESS MANUAL] Migración Histórica."
        if telefono_alternativo:
            texto_nota += f" | MDM: Intento de inyección con teléfono alternativo: {telefono_alternativo}"

        notas_historicas = {
            "notas": [{
                "tipo": "sistema",
                "contenido": texto_nota,
                "fecha": localtime(now()).isoformat(),
                "usuario": request.user.id
            }],
            "columnas_excel_historicas": {}
        }

        # --- 3. INYECCIÓN O FUSIÓN SILENCIOSA ---
        try:
            if tipo_entidad == 'CORPORATIVO':
                # El MDM siempre devuelve un objeto Clínica (nuevo o existente) para Corporativos
                CoreLead.objects.create(
                    owner=vendedor_asignado,
                    ubicacion=ubicacion_obj,
                    estatus='CLIENTE',  
                    phone_primary=telefono if not telefono_alternativo else instancia_mdm.telefono_master,
                    celular=celular[:15],
                    email=email,
                    nombre_pila=nombre_pila[:100],
                    clinica=instancia_mdm,
                    especialidad_cat=especialidad_obj,
                    producto_cat=producto_obj,
                    es_historico=True,
                    notas_variadas=notas_historicas,
                )
                messages.success(request, f"🚀 ¡Ingreso Corporativo exitoso! Vinculado a Clínica '{instancia_mdm.nombre}'. Asignado a {vendedor_asignado.username}.")

            else: # INDIVIDUAL
                if instancia_mdm is not None:
                    # FUSIÓN: El doctor ya estaba en BD, no creamos nuevo CoreLead.
                    if not isinstance(instancia_mdm.notas_variadas, dict):
                        instancia_mdm.notas_variadas = {"notas": [], "columnas_excel_historicas": {}}
                    if "notas" not in instancia_mdm.notas_variadas:
                        instancia_mdm.notas_variadas["notas"] = []
                        
                    nota_fusion = {
                        "tipo": "sistema",
                        "contenido": f"MDM Fusión: Inyección rápida interceptada. Registro validado. Teléfono extra insertado: {telefono_alternativo or telefono}",
                        "fecha": localtime(now()).isoformat(),
                        "usuario": request.user.id
                    }
                    instancia_mdm.notas_variadas["notas"].append(nota_fusion)
                    instancia_mdm.save()
                    messages.success(request, "Registro fusionado/actualizado exitosamente por MDM.")
                    
                else:
                    # NUEVO: Primer registro para este doctor, se crea desde cero.
                    CoreLead.objects.create(
                        owner=vendedor_asignado,
                        ubicacion=ubicacion_obj,
                        estatus='CLIENTE',  
                        phone_primary=telefono,
                        celular=celular[:15],
                        email=email,
                        titulo_cortesia=titulo_obj,
                        nombre_pila=nombre_pila[:100],
                        apellido_paterno=apellido_paterno[:100],
                        apellido_materno=apellido_materno[:100],
                        especialidad_cat=especialidad_obj,
                        producto_cat=producto_obj,
                        es_historico=True,
                        notas_variadas=notas_historicas,
                    )
                    messages.success(request, f"🚀 ¡Ingreso Express exitoso! Nuevo prospecto asignado a {vendedor_asignado.username}.")
            
            return redirect('ingesta_express')
            
        except Exception as e:
            messages.error(request, f"Error interno al guardar en BD: {str(e)}")
            return redirect('ingesta_express')

# 2. FICHA DE TRABAJO (Blindada contra el auto-formateador de VS Code)
@method_decorator(role_required(['VENDEDOR', 'DIRECTOR', 'ADMIN']), name='dispatch')
class FichaTrabajoView(LoginRequiredMixin, LeadOwnershipMixin, TemplateView):
    template_name = 'ficha_trabajo.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        lead_id = self.kwargs.get('pk') or self.kwargs.get('id')
        lead = get_object_or_404(CoreLead, id=lead_id)
        
        # Mandamos el objeto completo
        context['lead'] = lead
        # --- NUEVAS LÍNEAS PARA LOS DROPDOWNS ---
        context['especialidades_list'] = CatEspecialidad.objects.filter(is_active=True).values_list('nombre', flat=True).order_by('nombre')
        context['productos_list'] = CatProducto.objects.filter(is_active=True).order_by('nombre')
        context['titulos_list'] = CatTitulo.objects.filter(is_active=True).order_by('nombre')
        context['ubicaciones_list'] = CatUbicacion.objects.filter(is_active=True).values_list('ciudad', flat=True).order_by('ciudad')
        # Variables súper cortas para que el HTML no se rompa al guardar
        context['celular_seguro'] = lead.celular if lead.celular else "No registrado"
        # Priorizamos el catálogo relacional (DDS Fase 2)
        context['especialidad_segura'] = lead.especialidad_cat.nombre if lead.especialidad_cat else "No especificada"
        context['producto_seguro'] = lead.producto_cat.nombre if lead.producto_cat else "No especificado"

        # --- Validación flexible de estatus CLIENTE ---
        estatus_limpio = (lead.estatus or '').strip().upper()
        context['es_cliente'] = (estatus_limpio == 'CLIENTE')

        # --- Auto-crear TrackingPostVenta si el lead es CLIENTE ---
        if estatus_limpio == 'CLIENTE':
            from .models import TrackingPostVenta
            TrackingPostVenta.objects.get_or_create(lead=lead)
        
        return context

import json
from datetime import datetime
from datetime import timedelta
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required
from users.models import CatUbicacion
from django.contrib.auth import get_user_model

User = get_user_model()

import unicodedata
import re

from .mdm_services import normalizar_texto, limpiar_telefono_estricto, evaluar_duplicidad_estricta

def obtener_catalogos_limpios(texto_especialidad, texto_producto):
    """Versión Estricta DDS 2.0: Solo lee, NUNCA crea."""
    producto_obj = CatProducto.objects.filter(nombre__iexact=str(texto_producto).strip()).first()
    if not producto_obj:
        producto_obj = CatProducto.objects.filter(nombre__icontains='Por Definir').first()

    especialidad_obj = CatEspecialidad.objects.filter(nombre__iexact=str(texto_especialidad).strip()).first()
    if not especialidad_obj:
        especialidad_obj = CatEspecialidad.objects.filter(nombre__icontains='General').first()

    return especialidad_obj, producto_obj
    

    # 1. Procesar Producto (Estricto - Bóveda de 14 productos)
    texto_prod_norm = normalizar_texto(texto_producto)
    if texto_prod_norm:
        # Buscamos coincidencia exacta o en alias
        producto_obj = CatProducto.objects.filter(
            Q(nombre__iexact=texto_producto) | 
            Q(alias__icontains=texto_prod_norm)
        ).first()
        
        # Si no lo encuentra, asignamos el "Por Definir"
        if not producto_obj:
            producto_obj, _ = CatProducto.objects.get_or_create(nombre='Por Definir / Otro')

    # 2. Procesar Especialidad (Dinámico y Estético)
    texto_esp_limpio = str(texto_especialidad).strip()
    if not texto_esp_limpio:
        texto_esp_limpio = "General"
        
    # Buscamos si ya existe ignorando mayúsculas/minúsculas
    especialidad_obj = CatEspecialidad.objects.filter(nombre__iexact=texto_esp_limpio).first()
    
    # Si no existe, lo creamos respetando su formato bonito original
    if not especialidad_obj:
        especialidad_obj = CatEspecialidad.objects.create(nombre=texto_esp_limpio)

    return especialidad_obj, producto_obj

@login_required
@require_POST
def procesar_ingesta_masiva(request):
    if not request.user.is_authenticated:
        return JsonResponse({"error": "No autenticado"}, status=401)
    
    try:
        payload = json.loads(request.body)
        
        # Soporte para el nuevo formato con "action" o el antiguo directo
        if isinstance(payload, dict) and "action" in payload:
            action = payload.get("action")
            data = payload.get("leads", [])
        else:
            action = "commit"
            data = payload

        preview_results = []
        mapa_ciudades = {normalizar_texto(u.ciudad): u for u in CatUbicacion.objects.all()}

        from .parser_service import parsear_fila
        from leads.services.common_services import obtener_catalogos_limpios
        from .models import Clinica, LeadStaging

        inserted_count = 0

        for idx, lead_data in enumerate(data):
            if not isinstance(lead_data, dict):
                continue
            
            parsed_data = parsear_fila(lead_data)
            
            LeadStaging.objects.create(
                owner=request.user,
                datos_crudos=lead_data,
                datos_parseados=parsed_data,
                motivo_conflicto="Ingesta Masiva Vendedor (Curación Obligatoria)",
                estatus='PENDIENTE',
                origen='AGENTE'
            )
            inserted_count += 1

        return JsonResponse({'status': 'success', 'inserted': inserted_count})

    except json.JSONDecodeError:
        return JsonResponse({"error": "JSON Inválido enviado en la petición."}, status=400)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({"error": str(e)}, status=500)


@login_required
@require_POST
def procesar_alta_manual(request):
    if not request.user.is_authenticated:
        return JsonResponse({"error": "No autenticado"}, status=401)
    
    try:
        data = json.loads(request.body)
        from leads.services.lead_creation_service import crear_prospecto_core
        
        # Delegamos TODA la lógica al Service Layer
        resultado = crear_prospecto_core(data, request.user)
        
        if resultado.get("success"):
            return JsonResponse({
                'status': 'success', 
                'mensaje': resultado.get("mensaje", "Creado con éxito"),
                'lead_id': resultado.get("lead_id")
            })
        else:
            return JsonResponse({"error": resultado.get("error")}, status=resultado.get("status_code", 400))

    except json.JSONDecodeError:
        return JsonResponse({"error": "Datos inválidos enviados desde el formulario."}, status=400)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@login_required
@require_POST
def actualizar_lead_fsm(request, pk):
    if not request.user.is_authenticated:
        return JsonResponse({"error": "No autenticado"}, status=401)
    
    try:
        data = json.loads(request.body)
        from leads.services.fsm_services import procesar_transicion_fsm
        
        resultado = procesar_transicion_fsm(pk, data, request.user)
        
        if resultado.get("success"):
            # Auto-resolución: si el agente está tocando el lead, ya no está estancado
            from .models import Notificacion
            Notificacion.objects.filter(
                lead_id=pk, tipo__in=['estancamiento', 'reactivacion'], leida=False
            ).update(leida=True)

            response_data = {
                "status": resultado.get("status", "success"),
                "mensaje": resultado.get("mensaje"),
                "nuevo_estatus": resultado.get("nuevo_estatus")
            }
            if "datos_procesados" in resultado:
                response_data["datos_procesados"] = resultado["datos_procesados"]
            if "url_descarga" in resultado:
                response_data["url_descarga"] = resultado["url_descarga"]

            return JsonResponse(response_data)
        else:
            return JsonResponse({"error": resultado.get("error")}, status=resultado.get("status_code", 400))

    except json.JSONDecodeError:
        return JsonResponse({"error": "Datos inválidos enviados desde el formulario."}, status=400)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

@login_required
@require_POST
def api_marcar_no_cierre(request, pk):
    try:
        data = json.loads(request.body)
        motivo = data.get('motivo', 'Desconocido').strip()
        
        lead = get_object_or_404(CoreLead, id=pk)
        
        # Invocamos estrictamente a la máquina de estados FSM
        try:
            texto_nota_historica = f"Rechazo Definitivo: {motivo}"
            lead.archivar_sin_exito(nota_motivo=texto_nota_historica, usuario_id=request.user.id)
            lead.save()
            
            return JsonResponse({"status": "success", "mensaje": "El ciclo del prospecto se ha cerrado sin éxito (NO CIERRE)."})
            
        except Exception as e:
            # Capturamos excepciones propias del FSM (ej. TransitionNotAllowed)
            return JsonResponse({"error": str(e)}, status=400)
            
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


def es_director(user):
    return user.is_superuser


@login_required
@user_passes_test(es_director, login_url='dashboard_agente')
def director_dashboard_view(request):
    from leads.services.dashboard_services import obtener_metricas_director
    filtros = {
        'estado': request.GET.get('estado', ''),
        'especialidad': request.GET.get('especialidad', ''),
        'producto': request.GET.get('producto', ''),
        'vendedor': request.GET.get('vendedor', '')
    }
    context = obtener_metricas_director(filtros)
    return render(request, 'director_dashboard.html', context)

def marcar_alerta_leida(request, alerta_id):
    # Buscamos la alerta asegurándonos de que pertenezca a este usuario
    alerta = get_object_or_404(Notificacion, id=alerta_id, usuario=request.user)
    
    # La apagamos
    alerta.leida = True
    alerta.save()
    
    # Si la alerta está ligada a un prospecto, llevamos al vendedor a ese perfil
    if alerta.lead:
        # OJO: Cambia 'detalle_lead' por el nombre real de tu URL para ver un lead
        return redirect('detalle_lead', pk=alerta.lead.id) 
    
    # Si es una alerta general, lo regresamos al dashboard
    return redirect('dashboard_agente')


@login_required
@require_POST
def api_atender_alerta(request, alerta_id):
    """
    Endpoint para atender/descartar una alerta (Notificacion) mediante POST.
    - Marca la alerta como leída.
    - Inyecta auditoría en notas_variadas del CoreLead asociado.
    - Regla del Reloj Flotante: si el tipo es 'mantenimiento',
      crea un VentaTransaccional con estatus DESCARTADO para reiniciar el conteo de 18 meses.
    """
    from .models import Notificacion, VentaTransaccional
    try:
        data = json.loads(request.body)
        motivo = data.get('motivo', '').strip()

        alerta = get_object_or_404(Notificacion, id=alerta_id, usuario=request.user)

        # 1. Apagar la alerta
        alerta.leida = True
        alerta.save(update_fields=['leida'])

        # 2. Auditoría en notas_variadas del Lead asociado
        if alerta.lead:
            lead = alerta.lead
            tipo_display = alerta.get_tipo_display()
            timestamp = localtime(now()).strftime("%Y-%m-%d %H:%M")

            nueva_nota = {
                "fecha": timestamp,
                "tipo": "sistema",
                "contenido": f"🚨 Alerta de {tipo_display} atendida/descartada. Motivo: {motivo or 'Sin motivo registrado'}."
            }

            if not isinstance(lead.notas_variadas, dict):
                lead.notas_variadas = {"notas": []}
            if "notas" not in lead.notas_variadas:
                lead.notas_variadas["notas"] = []

            lead.notas_variadas["notas"].append(nueva_nota)
            lead.save(update_fields=['notas_variadas'])

            # 3. Regla del Reloj Flotante (solo para tipo 'mantenimiento')
            if alerta.tipo == 'mantenimiento':
                # Buscar o crear un producto genérico de familia SERVICIO
                producto_servicio, _ = CatProducto.objects.get_or_create(
                    familia='SERVICIO',
                    nombre='Servicio de mantenimiento',
                    defaults={'is_active': True}
                )
                VentaTransaccional.objects.create(
                    lead=lead,
                    producto=producto_servicio,
                    vendedor=request.user,
                    estatus='DESCARTADO',
                    notas=f"Rechazo desde Alerta (18 meses). Motivo: {motivo or 'Sin motivo registrado'}."
                )

        return JsonResponse({
            "status": "success",
            "mensaje": "Alerta atendida correctamente.",
            "lead_id": str(alerta.lead.id) if alerta.lead else None
        })

    except json.JSONDecodeError:
        return JsonResponse({"error": "JSON inválido."}, status=400)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

@login_required
@require_POST
def registrar_venta_extra(request):
    """
    Registra una compra transaccional solo si el lead ya es CLIENTE.
    Pensado para ventas de tipo Accesorio, Servicio o Evento.
    """
    try:
        data = json.loads(request.body)
        lead_id = data.get('lead_id')
        producto_id = data.get('producto_id')
        monto = data.get('monto')
        notas = data.get('notas', '')
        estatus = data.get('estatus', 'PENDIENTE') # Nuevo estatus

        if not lead_id or not producto_id:
            return JsonResponse({"error": "Faltan datos obligatorios (lead_id, producto_id)."}, status=400)
        
        lead = get_object_or_404(CoreLead, id=lead_id)
        
        # 1. Candado de Fidelización (Visión 360°): 
        # Cero impacto al FSM, venta reservada a CLIENTES ya consolidados en el embudo.
        if lead.estatus != 'CLIENTE':
            return JsonResponse({"error": "Solo puedes registrar ventas transaccionales a prospectos con estatus de CLIENTE."}, status=400)

        producto = get_object_or_404(CatProducto, id=producto_id)

        # --- CANDADO ANTI-SPAM (Código de antiG) ---
        from .models import VentaTransaccional
        
        if VentaTransaccional.objects.filter(lead=lead, producto=producto, estatus__in=['PENDIENTE', 'EN_GESTION']).exists():
            return JsonResponse({"error": "Ya existe una oportunidad activa (Pendiente o En Gestión) para este producto."}, status=400)
        # ------------------------------------------

        # 2. Guardado de Independencia Transaccional
        nueva_venta = VentaTransaccional.objects.create(
            lead=lead,
            producto=producto,
            vendedor=request.user,
            estatus=estatus, 
            monto=monto if monto else None,
            notas=notas
        )

        # --- NUEVO: Inyectar la miga de pan en el historial del Lead ---
        # ... (aquí sigue el resto de tu código normal que guarda las notas) ...
        # --- NUEVO: Inyectar la miga de pan en el historial del Lead ---
        # from django.utils import timezone
        
        #timestamp = timezone.now().strftime("%Y-%m-%d %H:%M")
        timestamp = localtime(now()).strftime("%Y-%m-%d %H:%M")
        
        nueva_nota = {
            "fecha": timestamp,
            "tipo": "sistema",  # Lo marcamos como sistema para que resalte
            "contenido": f"🎯 Oportunidad 360° creada: {producto.nombre} (Estatus: {estatus}).\n📝 Notas: {notas}"
        }

        # Asegurarnos de que el diccionario de notas exista
        if not isinstance(lead.notas_variadas, dict):
            lead.notas_variadas = {"notas": []}
        if "notas" not in lead.notas_variadas:
            lead.notas_variadas["notas"] = []
            
        # Guardar la nota y actualizar el lead
        lead.notas_variadas["notas"].append(nueva_nota)
        lead.save(update_fields=['notas_variadas'])
        # ----------------------------------------------------------------

        # --- Auto-resolución: si el producto es SERVICIO, apagar alertas de mantenimiento ---
        if producto.familia == 'SERVICIO':
            from .models import Notificacion
            Notificacion.objects.filter(
                lead=lead, tipo='mantenimiento', leida=False
            ).update(leida=True)

        return JsonResponse({
            "status": "success",
            "mensaje": "Venta transaccional registrada exitosamente.",
            "venta_id": str(nueva_venta.id)
        })

    except json.JSONDecodeError:
        return JsonResponse({"error": "JSON Inválido enviado en la petición."}, status=400)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

@login_required
def bandeja_rescate_view(request):
    if not request.user.is_superuser:
        return render(request, '403.html')
    
    leads = CoreLead.objects.filter(
        Q(estatus='EN_ESPERA') | Q(plan__iexact='descartado') | Q(estatus__iexact='NO_CIERRE')
    ).select_related('owner', 'especialidad_cat', 'producto_cat')
    
    vendedores = User.objects.filter(is_active=True, is_superuser=False)
    
    context = {
        'leads': leads,
        'vendedores': vendedores
    }
    return render(request, 'director_rescate.html', context)

@login_required
@require_POST
def api_reasignar_lead(request):
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'No autorizado'}, status=403)
        
    try:
        data = json.loads(request.body)
        lead_id = data.get('lead_id')
        nuevo_vendedor_id = data.get('nuevo_vendedor_id')
        
        lead = CoreLead.objects.get(id=lead_id)
        nuevo_vendedor = User.objects.get(id=nuevo_vendedor_id)
        
        lead.owner = nuevo_vendedor
        lead.estatus = 'LEAD'
        lead.plan = 'SEGUIMIENTO'
        lead.save()
        
        return JsonResponse({'success': True, 'message': 'Reasignado con éxito'})
        
    except CoreLead.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Lead no encontrado'}, status=404)
    except User.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Vendedor no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
@require_POST
def api_desechar_lead(request):
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'No autorizado'}, status=403)
        
    try:
        data = json.loads(request.body)
        lead_id = data.get('lead_id')
        
        lead = CoreLead.objects.get(id=lead_id)
        
        lead.plan = 'ARCHIVO_MUERTO'
        lead.save()
        
        return JsonResponse({'success': True, 'message': 'Lead desechado al archivo muerto'})
        
    except CoreLead.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Lead no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
def director_busqueda_view(request):
    if not request.user.is_superuser:
        return render(request, '403.html')
        
    q = request.GET.get('q', '').strip()
    leads = None
    
    if q:
        leads = CoreLead.objects.filter(
            Q(nombre_pila__icontains=q) |
            Q(apellido_paterno__icontains=q) |
            Q(apellido_materno__icontains=q) |
            Q(phone_primary__icontains=q) |
            Q(celular__icontains=q) |
            Q(email__icontains=q)
        ).select_related('owner', 'especialidad_cat', 'producto_cat')
    
    context = {
        'leads': leads,
        'q': q
    }
    
    return render(request, 'director_busqueda.html', context)

from django.core.paginator import Paginator

# ─── XLSX EXPORT HELPER ─────────────────────────────────────────────────────
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

_CAL_MAP = {3: 'Alta', 2: 'Media', 1: 'Baja'}

def generar_respuesta_xlsx(queryset, nombre_archivo):
    """
    Helper compartido. Recibe un QuerySet de CoreLead ya filtrado
    y devuelve un HttpResponse con el archivo .xlsx listo para descargar.
    """
    qs = queryset.select_related(
        'owner', 'especialidad_cat', 'producto_cat', 'ubicacion', 'titulo_cortesia'
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", fgColor="1E3A5F")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    HEADERS = [
        "Nombre Completo", "Telefono", "Celular", "Email",
        "Estatus", "Calificacion", "Producto", "Especialidad",
        "Ciudad", "Vendedor", "Fecha Registro", "Notas",
    ]
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
    ws.row_dimensions[1].height = 30

    for lead in qs:
        notas_raw  = lead.notas_variadas or {}
        notas_list = notas_raw.get("notas", []) if isinstance(notas_raw, dict) else []
        notas_texto = " | ".join(
            "[{}] {} -- {}".format(
                n.get('tipo', ''),
                n.get('contenido', ''),
                str(n.get('fecha', ''))[:10]
            )
            for n in notas_list if isinstance(n, dict)
        )

        ws.append([
            lead.nombre_completo_mdm,
            lead.phone_primary or "",
            lead.celular        or "",
            lead.email          or "",
            lead.estatus,
            _CAL_MAP.get(lead.calificacion, "Sin calificacion"),
            lead.producto_cat.nombre     if lead.producto_cat    else "",
            lead.especialidad_cat.nombre if lead.especialidad_cat else "",
            lead.ubicacion.ciudad        if lead.ubicacion        else "",
            lead.owner.username.title()  if lead.owner            else "",
            lead.created_at.strftime("%d/%m/%Y") if lead.created_at else "",
            notas_texto,
        ])

    ANCHOS = [32, 14, 14, 28, 18, 14, 24, 24, 18, 16, 16, 60]
    for i, ancho in enumerate(ANCHOS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = ancho

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="{}"'.format(nombre_archivo)
    wb.save(response)
    return response
# ─────────────────────────────────────────────────────────────────────────────


@login_required
def director_directorio_view(request):
    if not request.user.is_superuser:
        return render(request, '403.html')
        
    q = request.GET.get('q', '').strip()
    vendedor_id = request.GET.get('vendedor_id', '')
    estatus = request.GET.get('estatus', '')
    calificacion = request.GET.get('calificacion', '')
    
    leads = CoreLead.objects.all().select_related('owner', 'especialidad_cat', 'producto_cat')
    
    if q:
        leads = leads.filter(
            Q(nombre_pila__icontains=q) |
            Q(apellido_paterno__icontains=q) |
            Q(apellido_materno__icontains=q) |
            Q(phone_primary__icontains=q)
        )
    if vendedor_id:
        leads = leads.filter(owner_id=vendedor_id)
    if estatus:
        leads = leads.filter(estatus=estatus)
    if calificacion:
        leads = leads.filter(calificacion=int(calificacion))
        
    leads = leads.order_by('-created_at')
    
    paginator = Paginator(leads, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    vendedores = User.objects.filter(is_active=True, is_superuser=False)
    estatus_unicos = CoreLead.objects.values_list('estatus', flat=True).distinct().order_by('estatus')
    
    context = {
        'page_obj': page_obj,
        'q': q,
        'vendedor_id': vendedor_id,
        'estatus_filtro': estatus,
        'calificacion': calificacion,
        'vendedores': vendedores,
        'estatus_unicos': estatus_unicos,
    }
    
    return render(request, 'director_directorio.html', context)


@login_required
def director_directorio_exportar_view(request):
    """Exporta el directorio de leads a XLSX respetando los filtros activos."""
    if not request.user.is_superuser:
        return render(request, '403.html')

    q            = request.GET.get('q', '').strip()
    vendedor_id  = request.GET.get('vendedor_id', '')
    estatus      = request.GET.get('estatus', '')
    calificacion = request.GET.get('calificacion', '')

    leads = CoreLead.objects.all()

    if q:
        leads = leads.filter(
            Q(nombre_pila__icontains=q) |
            Q(apellido_paterno__icontains=q) |
            Q(apellido_materno__icontains=q) |
            Q(phone_primary__icontains=q)
        )
    if vendedor_id:
        leads = leads.filter(owner_id=vendedor_id)
    if estatus:
        leads = leads.filter(estatus=estatus)
    if calificacion:
        leads = leads.filter(calificacion=int(calificacion))

    leads = leads.order_by('-created_at')
    return generar_respuesta_xlsx(leads, "directorio_leads.xlsx")


from .models import Evento

@login_required
def director_eventos_view(request):
    if not request.user.is_superuser:
        return render(request, '403.html')
        
    ver_archivados = request.GET.get('ver_archivados') == '1'
    
    eventos_qs = Evento.objects.prefetch_related('vendedores_asignados', 'ciudades_objetivo')
    if not ver_archivados:
        eventos_qs = eventos_qs.exclude(estatus='ARCHIVADO')
        
    eventos = eventos_qs.order_by('-fecha_inicio')
    
    # Calcular estadísticas de prospectos y leads por evento
    for ev in eventos:
        ev.total_registros = ev.clientes_vinculados.count()
        ev.registros_lead = ev.clientes_vinculados.filter(lead__estatus='LEAD').count()
        ev.registros_calificado = ev.clientes_vinculados.filter(lead__estatus='LEAD_CALIFICADO').count()
        ev.registros_ventas = ev.clientes_vinculados.filter(lead__estatus='CLIENTE').count()
        
    vendedores = User.objects.filter(is_active=True, is_superuser=False).order_by('username')
    ciudades_list = CatUbicacion.objects.filter(is_active=True).order_by('estado', 'ciudad')
    
    context = {
        'eventos': eventos,
        'vendedores': vendedores,
        'ciudades_list': ciudades_list,
        'ver_archivados': ver_archivados,
    }
    return render(request, 'director_eventos.html', context)

@login_required
@require_POST
def api_crear_evento(request):
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'No autorizado'}, status=403)
        
    try:
        data = json.loads(request.body)
        
        nombre = data.get('nombre')
        tipo = data.get('tipo', 'EXPO')
        fecha_inicio = data.get('fecha_inicio')
        fecha_fin = data.get('fecha_fin')
        lugar = data.get('lugar')
        linea_producto = data.get('linea_producto', 'TODAS')
        ciudades_objetivo_ids = data.get('ciudades_objetivo', [])
        
        if not all([nombre, fecha_inicio, fecha_fin]):
            return JsonResponse({'success': False, 'error': 'Faltan campos obligatorios.'}, status=400)
            
        if tipo != 'CAMPAÑA' and not lugar:
            return JsonResponse({'success': False, 'error': 'El lugar es obligatorio para Expos y Talleres.'}, status=400)
            
        if tipo == 'CAMPAÑA':
            lugar = None
             
        vendedores_ids = data.get('vendedores_ids', [])
        
        nuevo_evento = Evento.objects.create(
            nombre=nombre,
            tipo=tipo,
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin,
            lugar=lugar,
            linea_producto=linea_producto
        )
        
        if ciudades_objetivo_ids:
            nuevo_evento.ciudades_objetivo.set(ciudades_objetivo_ids)
        
        if vendedores_ids:
            nuevo_evento.vendedores_asignados.set(vendedores_ids)
            
        return JsonResponse({'success': True, 'message': 'Evento creado correctamente.'})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
@require_POST
def api_eliminar_evento(request, evento_id):
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'No autorizado'}, status=403)
        
    try:
        evento = Evento.objects.get(id=evento_id)
        evento.delete()
        return JsonResponse({'success': True})
    except Evento.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Evento no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
@require_POST
def api_editar_evento(request):
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'No autorizado'}, status=403)
        
    try:
        data = json.loads(request.body)
        evento_id = data.get('id')
        
        evento = get_object_or_404(Evento, id=evento_id)
        
        nombre = data.get('nombre')
        tipo = data.get('tipo', 'EXPO')
        fecha_inicio = data.get('fecha_inicio')
        fecha_fin = data.get('fecha_fin')
        lugar = data.get('lugar')
        linea_producto = data.get('linea_producto', 'TODAS')
        estatus = data.get('estatus', 'ACTIVO')
        ciudades_objetivo_ids = data.get('ciudades_objetivo', [])
        vendedores_ids = data.get('vendedores_ids', [])
        
        if not all([nombre, fecha_inicio, fecha_fin]):
            return JsonResponse({'success': False, 'error': 'Faltan campos obligatorios.'}, status=400)
            
        if tipo != 'CAMPAÑA' and not lugar:
            return JsonResponse({'success': False, 'error': 'El lugar es obligatorio para Expos y Talleres.'}, status=400)
            
        if tipo == 'CAMPAÑA':
            lugar = None
            
        evento.nombre = nombre
        evento.tipo = tipo
        evento.fecha_inicio = fecha_inicio
        evento.fecha_fin = fecha_fin
        evento.lugar = lugar
        evento.linea_producto = linea_producto
        evento.estatus = estatus
        evento.save()
        
        evento.ciudades_objetivo.set(ciudades_objetivo_ids)
        evento.vendedores_asignados.set(vendedores_ids)
        
        return JsonResponse({'success': True})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_POST
def api_archivar_evento(request, evento_id):
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'No autorizado'}, status=403)
        
    try:
        evento = get_object_or_404(Evento, id=evento_id)
        
        # Leer la acción del body si viene como JSON
        action = 'archive'
        if request.body:
            try:
                data = json.loads(request.body)
                action = data.get('action', 'archive')
            except json.JSONDecodeError:
                pass
                
        if action == 'unarchive':
            from datetime import date
            if evento.fecha_fin < date.today():
                evento.estatus = 'FINALIZADO'
            else:
                evento.estatus = 'ACTIVO'
        else:
            evento.estatus = 'ARCHIVADO'
            
        evento.save()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def api_detalle_evento(request, evento_id):
    if not request.user.is_superuser:
        return JsonResponse({'success': False, 'error': 'No autorizado'}, status=403)
        
    try:
        evento = get_object_or_404(Evento, id=evento_id)
        
        # Obtener los clientes vinculados
        vinculaciones = evento.clientes_vinculados.select_related('lead', 'lead__owner', 'lead__especialidad_cat').order_by('-fecha_vinculacion')
        
        clientes = []
        vendedores_stats = {}
        
        # Inicializar vendedores asignados en las estadísticas
        for v in evento.vendedores_asignados.all():
            nombre_vendedor = f"{v.first_name} {v.last_name}".strip() or v.username
            vendedores_stats[v.id] = {
                'nombre': nombre_vendedor.title(),
                'cantidad': 0,
                'username': v.username
            }
            
        for vinc in vinculaciones:
            lead = vinc.lead
            vendedor = lead.owner
            
            # Nombre completo del vendedor
            vendedor_name = f"{vendedor.first_name} {vendedor.last_name}".strip() or vendedor.username
            vendedor_name = vendedor_name.title()
            
            # Registrar/incrementar en estadísticas
            if vendedor.id not in vendedores_stats:
                vendedores_stats[vendedor.id] = {
                    'nombre': vendedor_name,
                    'cantidad': 0,
                    'username': vendedor.username
                }
            vendedores_stats[vendedor.id]['cantidad'] += 1
            
            clientes.append({
                'lead_id': str(lead.id),
                'nombre': lead.nombre_completo_mdm,
                'especialidad': lead.especialidad_cat.nombre if lead.especialidad_cat else '-',
                'vendedor': vendedor_name,
                'fecha_vinculacion': vinc.fecha_vinculacion.strftime('%d/%m/%Y'),
                'comentarios': vinc.comentarios or ''
            })
            
        # Convertir estadísticas de vendedores a una lista ordenada por cantidad descendente
        vendedores_lista = sorted(list(vendedores_stats.values()), key=lambda x: x['cantidad'], reverse=True)
        
        return JsonResponse({
            'success': True,
            'evento': {
                'id': evento.id,
                'nombre': evento.nombre,
                'tipo': evento.tipo,
                'tipo_display': evento.get_tipo_display(),
                'linea_producto': evento.get_linea_producto_display(),
                'lugar': evento.lugar or '',
                'fecha_inicio': evento.fecha_inicio.strftime('%d/%m/%Y'),
                'fecha_fin': evento.fecha_fin.strftime('%d/%m/%Y'),
            },
            'total_registros': len(clientes),
            'vendedores_stats': vendedores_lista,
            'clientes': clientes
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_POST
def actualizar_estatus_venta_extra(request, venta_id):
    """
    Endpoint para actualizar el estatus de una VentaTransaccional (Oportunidad 360°).
    Inyecta una miga de pan en notas_variadas del CoreLead asociado.
    """
    try:
        data = json.loads(request.body)
        nuevo_estatus = data.get('estatus')

        # Validación de estatus permitido
        from .models import VentaTransaccional
        estatus_validos = [choice[0] for choice in VentaTransaccional.ESTATUS_CHOICES]
        if nuevo_estatus not in estatus_validos:
            return JsonResponse({"error": f"Estatus '{nuevo_estatus}' no es válido."}, status=400)

        venta = get_object_or_404(VentaTransaccional, id=venta_id)

        # --- Guardia de Flujo Unidireccional ---
        TRANSICIONES_PERMITIDAS = {
            'PENDIENTE':   ['EN_GESTION'],
            'EN_GESTION':  ['CONCRETADO', 'DESCARTADO'],
            'CONCRETADO':  [],   # Estado terminal
            'DESCARTADO':  [],   # Estado terminal
        }

        transiciones_validas = TRANSICIONES_PERMITIDAS.get(venta.estatus, [])
        if nuevo_estatus not in transiciones_validas:
            return JsonResponse({
                "error": f"Transición no permitida: {venta.get_estatus_display()} → {nuevo_estatus}. "
                         f"Solo puedes avanzar a: {', '.join(transiciones_validas) or 'Ninguno (estado final)'}."
            }, status=400)

        estatus_anterior = venta.get_estatus_display()

        # Actualizar el estatus
        venta.estatus = nuevo_estatus
        venta.save(update_fields=['estatus'])

        # --- Inyectar miga de pan en el historial del Lead ---
        lead = venta.lead
        timestamp = localtime(now()).strftime("%Y-%m-%d %H:%M")

        nueva_nota = {
            "fecha": timestamp,
            "tipo": "sistema",
            "contenido": f"🔄 Oportunidad 360° actualizada: {venta.producto.nombre} — {estatus_anterior} → {venta.get_estatus_display()}."
        }

        if not isinstance(lead.notas_variadas, dict):
            lead.notas_variadas = {"notas": []}
        if "notas" not in lead.notas_variadas:
            lead.notas_variadas["notas"] = []

        lead.notas_variadas["notas"].append(nueva_nota)

        # --- Nota opcional del vendedor ---
        nota_vendedor = data.get('nota', '').strip()
        if nota_vendedor:
            nota_humana = {
                "fecha": timestamp,
                "tipo": "vendedor",
                "contenido": f"📝 [{venta.producto.nombre}] {nota_vendedor}"
            }
            lead.notas_variadas["notas"].append(nota_humana)

        lead.save(update_fields=['notas_variadas'])

        return JsonResponse({
            "status": "success",
            "mensaje": f"Estatus actualizado a {venta.get_estatus_display()}."
        })

    except json.JSONDecodeError:
        return JsonResponse({"error": "JSON inválido."}, status=400)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@login_required
@user_passes_test(es_director, login_url='dashboard_agente')
def dashboard_fidelizacion_view(request):
    """
    Dashboard de Fidelización 360° — Métricas de volumen
    para oportunidades de venta cruzada (VentaTransaccional).
    Solo accesible para la Dirección.
    """
    from .models import VentaTransaccional
    from django.core.paginator import Paginator
    from datetime import datetime

    # 1. Filtros
    filtro_vendedor = request.GET.get('vendedor', '')
    filtro_estatus = request.GET.get('estatus', '')
    filtro_familia = request.GET.get('familia', '')
    filtro_fecha_desde = request.GET.get('fecha_desde', '')
    filtro_fecha_hasta = request.GET.get('fecha_hasta', '')

    qs = VentaTransaccional.objects.select_related('lead', 'producto', 'vendedor').all()

    if filtro_vendedor:
        qs = qs.filter(vendedor__username__iexact=filtro_vendedor)
    if filtro_estatus:
        qs = qs.filter(estatus=filtro_estatus)
    if filtro_familia:
        qs = qs.filter(producto__familia__iexact=filtro_familia)

    # Filtros de rango de fechas (sobre fecha_venta)
    if filtro_fecha_desde:
        try:
            fecha_desde = datetime.strptime(filtro_fecha_desde, '%Y-%m-%d')
            qs = qs.filter(fecha_venta__date__gte=fecha_desde.date())
        except ValueError:
            pass
    if filtro_fecha_hasta:
        try:
            fecha_hasta = datetime.strptime(filtro_fecha_hasta, '%Y-%m-%d')
            qs = qs.filter(fecha_venta__date__lte=fecha_hasta.date())
        except ValueError:
            pass

    qs = qs.order_by('-fecha_venta')

    # 2. KPIs (Volumen y Financieros)
    total_ops = qs.count()
    pendientes = qs.filter(estatus='PENDIENTE').count()
    concretadas = qs.filter(estatus='CONCRETADO').count()
    en_gestion = qs.filter(estatus='EN_GESTION').count()
    descartadas = qs.filter(estatus='DESCARTADO').count()

    tasa_cierre = 0
    ops_cerradas = concretadas + descartadas
    if ops_cerradas > 0:
        tasa_cierre = round((concretadas / ops_cerradas) * 100, 1)

    from django.db.models import Sum, Avg
    
    # Facturación y Ticket Promedio
    monto_concretado = qs.filter(estatus='CONCRETADO').aggregate(total=Sum('monto'))['total'] or 0
    monto_concretado = round(monto_concretado, 2)
    
    ticket_promedio = qs.filter(estatus='CONCRETADO').aggregate(promedio=Avg('monto'))['promedio'] or 0
    ticket_promedio = round(ticket_promedio, 2)

    # Distribución por familia (concretadas)
    acc_count = qs.filter(estatus='CONCRETADO', producto__familia='ACCESORIO').count()
    ser_count = qs.filter(estatus='CONCRETADO', producto__familia='SERVICIO').count()
    eve_count = qs.filter(estatus='CONCRETADO', producto__familia='EVENTO').count()
    
    chart_familia_data = json.dumps({
        'labels': ['Accesorios', 'Servicios', 'Eventos'],
        'data': [acc_count, ser_count, eve_count],
        'colors': ['#ec4899', '#06b6d4', '#ffc107'], # Pink, Teal, Yellow
    })

    # Atribución por eventos (Talleres y Campañas)
    # 1. Ventas concretadas influenciadas por Talleres
    ventas_taller = qs.filter(
        estatus='CONCRETADO',
        lead__eventos_asociados__evento__tipo='TALLER'
    ).distinct()
    monto_taller = ventas_taller.aggregate(total=Sum('monto'))['total'] or 0
    monto_taller = round(monto_taller, 2)
    cantidad_taller = ventas_taller.count()

    # 2. Ventas concretadas influenciadas por Campañas
    ventas_campana = qs.filter(
        estatus='CONCRETADO',
        lead__eventos_asociados__evento__tipo='CAMPAÑA'
    ).distinct()
    monto_campana = ventas_campana.aggregate(total=Sum('monto'))['total'] or 0
    monto_campana = round(monto_campana, 2)
    cantidad_campana = ventas_campana.count()

    # 3. Datos para Chart.js (JSON seguro para el template)
    chart_data = json.dumps({
        'labels': ['Pendiente', 'En Gestión', 'Concretado', 'Descartado'],
        'data': [pendientes, en_gestion, concretadas, descartadas],
        'colors': ['#ffc107', '#0d6efd', '#198754', '#dc3545'],
    })

    # 4. Listas para Dropdowns
    vendedores_list = User.objects.filter(is_active=True, is_superuser=False).order_by('username')
    familias_list = ['ACCESORIO', 'SERVICIO', 'EVENTO']
    estatus_list = VentaTransaccional.ESTATUS_CHOICES

    # 5. Paginación (5 registros para convivir con la gráfica)
    paginator = Paginator(qs, 5)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'total_ops': total_ops,
        'pendientes': pendientes,
        'concretadas': concretadas,
        'en_gestion': en_gestion,
        'descartadas': descartadas,
        'tasa_cierre': tasa_cierre,
        'monto_concretado': monto_concretado,
        'ticket_promedio': ticket_promedio,
        'chart_data': chart_data,
        'chart_familia_data': chart_familia_data,
        'monto_taller': monto_taller,
        'cantidad_taller': cantidad_taller,
        'monto_campana': monto_campana,
        'cantidad_campana': cantidad_campana,
        'vendedores': vendedores_list,
        'familias': familias_list,
        'estatus_list': estatus_list,
        'filtro_vendedor': filtro_vendedor,
        'filtro_estatus': filtro_estatus,
        'filtro_familia': filtro_familia,
        'filtro_fecha_desde': filtro_fecha_desde,
        'filtro_fecha_hasta': filtro_fecha_hasta,
    }
    return render(request, 'director_fidelizacion.html', context)


@login_required
@require_POST
def api_marcar_hito_postventa(request, lead_id):
    """
    Marca un hito post-venta (capacitacion o calidad) como completado.
    Actualiza TrackingPostVenta e inyecta miga de pan en notas_variadas.
    """
    from .models import TrackingPostVenta
    try:
        data = json.loads(request.body)
        hito = data.get('hito', '')

        lead = get_object_or_404(CoreLead, id=lead_id)
        tracking, _ = TrackingPostVenta.objects.get_or_create(lead=lead)

        timestamp = localtime(now()).strftime("%Y-%m-%d %H:%M")

        if hito == 'capacitacion':
            if tracking.capacitacion_dada:
                return JsonResponse({"error": "La capacitación ya fue marcada como realizada."}, status=400)
            tracking.capacitacion_dada = True
            tracking.fecha_capacitacion = now()
            tracking.save(update_fields=['capacitacion_dada', 'fecha_capacitacion'])
            emoji = "🎓"
            nombre_hito = "Capacitación Post-Venta"
            # Auto-resolución: apagar alertas vivas de tipo 'capacitacion' para este lead
            from .models import Notificacion
            Notificacion.objects.filter(
                lead=lead, tipo='capacitacion', leida=False
            ).update(leida=True)

        elif hito == 'calidad':
            if tracking.calidad_hecha:
                return JsonResponse({"error": "La llamada de calidad ya fue marcada como realizada."}, status=400)
            tracking.calidad_hecha = True
            tracking.fecha_calidad = now()
            tracking.save(update_fields=['calidad_hecha', 'fecha_calidad'])
            emoji = "📞"
            nombre_hito = "Llamada de Calidad"
            # Auto-resolución: apagar alertas vivas de tipo 'calidad' para este lead
            from .models import Notificacion
            Notificacion.objects.filter(
                lead=lead, tipo='calidad', leida=False
            ).update(leida=True)

        else:
            return JsonResponse({"error": "Hito no válido. Use 'capacitacion' o 'calidad'."}, status=400)

        # --- Inyectar miga de pan ---
        nueva_nota = {
            "fecha": timestamp,
            "tipo": "sistema",
            "contenido": f"{emoji} Hito Post-Venta completado: {nombre_hito}."
        }

        if not isinstance(lead.notas_variadas, dict):
            lead.notas_variadas = {"notas": []}
        if "notas" not in lead.notas_variadas:
            lead.notas_variadas["notas"] = []

        lead.notas_variadas["notas"].append(nueva_nota)
        lead.save(update_fields=['notas_variadas'])

        return JsonResponse({
            "status": "success",
            "mensaje": f"✅ {nombre_hito} marcado como completado."
        })

    except json.JSONDecodeError:
        return JsonResponse({"error": "JSON inválido."}, status=400)
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


class AgenteStagingListView(LoginRequiredMixin, ListView):
    model = LeadStaging
    template_name = 'leads/agente_staging_list.html'
    context_object_name = 'leads_staging'
    paginate_by = 50

    def get_queryset(self):
        return LeadStaging.objects.filter(owner=self.request.user, estatus='PENDIENTE', origen='AGENTE').order_by('created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['total_pendientes'] = LeadStaging.objects.filter(owner=self.request.user, estatus='PENDIENTE', origen='AGENTE').count()
        return context

class AgenteStagingProcesarView(LoginRequiredMixin, DetailView):
    model = LeadStaging
    template_name = 'leads/agente_staging_procesar.html'
    context_object_name = 'staging_lead'

    def get_queryset(self):
        return LeadStaging.objects.filter(owner=self.request.user, estatus='PENDIENTE', origen='AGENTE')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['especialidades_list'] = CatEspecialidad.objects.filter(is_active=True).order_by('nombre')
        context['ubicaciones_list'] = CatUbicacion.objects.filter(is_active=True).order_by('ciudad')
        context['titulos_list'] = CatTitulo.objects.filter(is_active=True).order_by('nombre')
        context['productos_list'] = CatProducto.objects.filter(is_active=True, familia='EQUIPO').order_by('nombre')
        context['restantes'] = LeadStaging.objects.filter(owner=self.request.user, estatus='PENDIENTE', origen='AGENTE').count()
        return context

    def post(self, request, *args, **kwargs):
        staging_lead = self.get_object()
        action = request.POST.get('action')

        if action == 'descartar':
            staging_lead.estatus = 'DESCARTADO'
            staging_lead.save()
            messages.info(request, "Registro descartado y eliminado de tu Quirófano.")
            return redirect('agente_staging_list')

        elif action == 'guardar':
            tipo_entidad = request.POST.get('tipo_entidad', 'INDIVIDUAL')
            especialidad_id = request.POST.get('especialidad_id')
            ubicacion_id = request.POST.get('ubicacion_id')
            producto_id = request.POST.get('producto_id')

            telefono = str(request.POST.get('telefono', '')).strip()
            celular = str(request.POST.get('celular', '')).strip()
            email = str(request.POST.get('email', '')).strip()
            direccion = str(request.POST.get('direccion_completa', '')).strip()

            titulo_id = request.POST.get('titulo_cortesia')
            nombre_pila = str(request.POST.get('nombre_pila', '')).strip()
            apellido_paterno = str(request.POST.get('apellido_paterno', '')).strip()
            apellido_materno = str(request.POST.get('apellido_materno', '')).strip()

            notas_originales = str(request.POST.get('notas', staging_lead.datos_crudos.get('notas', ''))).strip()

            titulo_obj = CatTitulo.objects.filter(id=titulo_id).first() if titulo_id else None
            especialidad_obj = CatEspecialidad.objects.filter(id=especialidad_id).first() if especialidad_id else None
            ubicacion_obj = CatUbicacion.objects.filter(id=ubicacion_id).first() if ubicacion_id else None
            producto_obj = CatProducto.objects.filter(id=producto_id).first() if producto_id else None

            if not all([especialidad_obj, ubicacion_obj]):
                messages.error(request, "Especialidad y Ubicación son campos obligatorios.")
                return redirect('agente_staging_procesar', pk=staging_lead.pk)

            datos_dict = {
                'tipo_entidad': tipo_entidad, 'nombre_pila': nombre_pila,
                'apellido_paterno': apellido_paterno, 'apellido_materno': apellido_materno,
                'telefono': telefono, 'especialidad_obj': especialidad_obj, 'ubicacion_obj': ubicacion_obj
            }

            try:
                from leads.services.mdm_service import resolver_identidad
                instancia_mdm, telefono_alternativo = resolver_identidad(datos_dict)
            except ValueError as e:
                messages.error(request, str(e))
                return redirect('agente_staging_procesar', pk=staging_lead.pk)

            texto_nota = f"[QUIRÓFANO] Validado por Vendedor. Notas: {notas_originales}"
            if telefono_alternativo:
                texto_nota += f" | MDM: Inyección con teléfono alternativo: {telefono_alternativo}"

            notas_historicas = {
                "notas": [{"tipo": "sistema", "contenido": texto_nota, "fecha": localtime(now()).isoformat(), "usuario": request.user.id}],
                "columnas_excel_crudas": staging_lead.datos_crudos
            }

            try:
                if tipo_entidad == 'CORPORATIVO':
                    from .models import Clinica
                    CoreLead.objects.create(
                        owner=request.user, ubicacion=ubicacion_obj, estatus='PROSPECTO',  
                        phone_primary=telefono if not telefono_alternativo else instancia_mdm.telefono_master,
                        celular=celular[:15], email=email, direccion_completa=direccion[:255],
                        nombre_pila=nombre_pila[:100], clinica=instancia_mdm,
                        especialidad_cat=especialidad_obj, producto_cat=producto_obj,
                        notas_variadas=notas_historicas,
                    )
                    messages.success(request, f"¡Agregado exitosamente a Corporativo '{instancia_mdm.nombre_clinica if hasattr(instancia_mdm, 'nombre_clinica') else instancia_mdm.nombre}'!")
                else: 
                    if instancia_mdm is not None:
                        if not isinstance(instancia_mdm.notas_variadas, dict):
                            instancia_mdm.notas_variadas = {"notas": [], "columnas_excel_crudas": {}}
                        if "notas" not in instancia_mdm.notas_variadas:
                            instancia_mdm.notas_variadas["notas"] = []
                        
                        instancia_mdm.notas_variadas["notas"].append({
                            "tipo": "sistema", "contenido": f"MDM Quirófano Agente: Fusión confirmada. Notas: {notas_originales}",
                            "fecha": localtime(now()).isoformat(), "usuario": request.user.id
                        })
                        instancia_mdm.save()
                        messages.success(request, f"Prospecto fusionado con el registro MDM existente.")
                    else:
                        CoreLead.objects.create(
                            owner=request.user, ubicacion=ubicacion_obj, estatus='PROSPECTO',
                            phone_primary=telefono if not telefono_alternativo else telefono_alternativo,
                            celular=celular[:15], email=email, direccion_completa=direccion[:255],
                            titulo_cortesia=titulo_obj, nombre_pila=nombre_pila[:100],
                            apellido_paterno=apellido_paterno[:100], apellido_materno=apellido_materno[:100],
                            especialidad_cat=especialidad_obj, producto_cat=producto_obj,
                            notas_variadas=notas_historicas,
                        )
                        messages.success(request, f"Lead Inyectado Formalmente de Quirófano a tu Pipeline Transaccional.")

                staging_lead.estatus = 'RESUELTO'
                staging_lead.save()
                return redirect('agente_staging_list')

            except Exception as e:
                messages.error(request, f"Error de DB: {str(e)}")
                return redirect('agente_staging_procesar', pk=staging_lead.pk)

from django.http import JsonResponse

@login_required
def api_citas_dia(request):
    fecha_str = request.GET.get('fecha')
    if not fecha_str:
        return JsonResponse({'error': 'Fecha no proporcionada'}, status=400)
    
    try:
        from django.utils.dateparse import parse_date
        fecha = parse_date(fecha_str)
        if not fecha:
            return JsonResponse({'error': 'Formato de fecha inválido'}, status=400)
            
        citas_count = CoreLead.objects.filter(
            owner=request.user,
            next_action_date=fecha
        ).count()
        
        return JsonResponse({'citas_programadas': citas_count})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_POST
def vincular_cliente_evento_view(request):
    """
    Endpoint para vincular o desvincular un cliente (estatus CLIENTE únicamente)
    a/de un evento (Taller o Campaña) activo asignado al vendedor.
    """
    try:
        data = json.loads(request.body)
        lead_id = data.get('lead_id')
        evento_id = data.get('evento_id')
        accion = data.get('accion') # 'vincular' o 'desvincular'

        if not all([lead_id, evento_id, accion]):
            return JsonResponse({'success': False, 'error': 'Faltan parámetros obligatorios.'}, status=400)

        if accion not in ['vincular', 'desvincular']:
            return JsonResponse({'success': False, 'error': 'Acción no válida.'}, status=400)

        from .models import CoreLead, Evento, LeadEvento
        lead = get_object_or_404(CoreLead, id=lead_id)

        # 1. Validación de propiedad
        if lead.owner != request.user:
            return JsonResponse({'success': False, 'error': 'No tienes permisos sobre este cliente.'}, status=403)

        # 2. Validación de estatus CLIENTE únicamente
        if lead.estatus != 'CLIENTE':
            return JsonResponse({'success': False, 'error': 'Este evento solo está disponible para registros con estatus CLIENTE.'}, status=400)

        # 3. Validación del Evento
        evento = get_object_or_404(Evento, id=evento_id)
        if evento.estatus != 'ACTIVO':
            return JsonResponse({'success': False, 'error': 'El evento no está activo.'}, status=400)
        
        if evento.tipo not in ['TALLER', 'CAMPAÑA']:
            return JsonResponse({'success': False, 'error': 'Tipo de evento no permitido.'}, status=400)

        if not evento.vendedores_asignados.filter(id=request.user.id).exists():
            return JsonResponse({'success': False, 'error': 'No estás asignado a este evento.'}, status=403)

        # 4. Operación
        if accion == 'vincular':
            obj, created = LeadEvento.objects.get_or_create(evento=evento, lead=lead)
            mensaje = 'Cliente vinculado con éxito.'
        else:
            LeadEvento.objects.filter(evento=evento, lead=lead).delete()
            mensaje = 'Cliente desvinculado con éxito.'

        return JsonResponse({'success': True, 'mensaje': mensaje})

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def obtener_eventos_cliente_view(request):
    """
    Retorna la lista de talleres y campañas activos del vendedor y si el cliente
    (estatus CLIENTE únicamente) está vinculado a ellos.
    Filtra los eventos según la compatibilidad dinámica con la especialidad del cliente.
    """
    lead_id = request.GET.get('lead_id')
    if not lead_id:
        return JsonResponse({'success': False, 'error': 'Falta el parámetro lead_id.'}, status=400)

    from .models import CoreLead, Evento, LeadEvento
    lead = get_object_or_404(CoreLead, id=lead_id)

    if lead.owner != request.user:
        return JsonResponse({'success': False, 'error': 'No tienes permisos sobre este cliente.'}, status=403)

    if lead.estatus != 'CLIENTE':
        return JsonResponse({'success': False, 'error': 'Este cliente no es de estatus CLIENTE.'}, status=400)

    # Eventos activos asignados al vendedor (excluyendo EXPO)
    eventos = Evento.objects.filter(vendedores_asignados=request.user, estatus='ACTIVO').exclude(tipo='EXPO').order_by('nombre')
    
    # IDs de eventos a los que el cliente ya está vinculado
    eventos_vinculados_ids = set(LeadEvento.objects.filter(lead=lead).values_list('evento_id', flat=True))

    eventos_data = []
    for ev in eventos:
        # Filtrar si la especialidad del cliente está permitida para la línea del evento,
        # o si el cliente ya está vinculado a él (permitiendo desvinculación)
        ids_permitidos = obtener_especialidades_permitidas(ev.linea_producto)
        if (lead.especialidad_cat_id in ids_permitidos) or (ev.id in eventos_vinculados_ids):
            eventos_data.append({
                'id': ev.id,
                'nombre': ev.nombre,
                'tipo': ev.tipo,
                'tipo_display': ev.get_tipo_display(),
                'fecha_inicio': ev.fecha_inicio.strftime('%d/%m/%Y'),
                'fecha_fin': ev.fecha_fin.strftime('%d/%m/%Y'),
                'vinculado': ev.id in eventos_vinculados_ids
            })

    return JsonResponse({'success': True, 'eventos': eventos_data})


def obtener_especialidades_permitidas(linea_producto):
    """
    Retorna la lista de IDs de especialidades permitidas para una línea de producto,
    combinando las reglas por defecto con las excepciones registradas en la base de datos.
    """
    from users.models import CatEspecialidad
    from .models import ExcepcionEspecialidadLinea
    
    # 1. Obtener todas las especialidades activas
    todas_especialidades = list(CatEspecialidad.objects.filter(is_active=True))
    
    # 2. Mapeo de reglas por defecto
    MAP_LINEA_DATOS = {
        'SPORT': ['Fisioterapeuta', 'Medicina del Deporte', 'Acupuntorista', 'Alfabiotismo', 'Angiólogo', 'Quiropráctica', 'Ortopedista', 'Homeópata'],
        'PET': ['Veterinario'],
        'DENTAL': ['Cirujano Dentista', 'Dentista', 'Odontólogo'],
        'PODOLOGICO': ['Podólogo'],
        'BEAUTY': ['Cirujano Plástico', 'Cosmetóloga', 'Dermatólogo'],
    }
    especialidades_defecto = MAP_LINEA_DATOS.get(linea_producto, [])
    
    # 3. Obtener excepciones explícitas para esta línea de producto
    excepciones = {exc.especialidad_id: exc.permitido for exc in ExcepcionEspecialidadLinea.objects.filter(linea_producto=linea_producto)}
    
    # 4. Construir la lista de IDs permitidos
    ids_permitidos = []
    for esp in todas_especialidades:
        # Si tiene una excepción explícita, se usa el valor de la excepción (permitido True/False)
        if esp.id in excepciones:
            if excepciones[esp.id]:
                ids_permitidos.append(esp.id)
        # Si no tiene excepción, se usa la regla por defecto
        else:
            # Si la línea no tiene restricciones por defecto (ej. TODAS, SERVICIO, ACCESORIO), se permite
            if not especialidades_defecto or esp.nombre in especialidades_defecto:
                ids_permitidos.append(esp.id)
                
    return ids_permitidos



